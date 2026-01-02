
import openpyxl
import os
import glob
import json
import re
from datetime import datetime

def parse_excel_folder(folder_path):
    all_data = []
    files = glob.glob(os.path.join(folder_path, "勤務*.xlsx"))
    files.sort()
    
    for file_path in files:
        print(f"Processing: {os.path.basename(file_path)}")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                user_name = sheet_name.strip()
                if not user_name or user_name in ["合計", "Sheet1", "マスタ", "基本形", "1", "2"]: continue
                
                sheet = wb[sheet_name]
                
                year = 2024
                month = 4
                
                # ヘッダー情報の取得を強化
                # B2, C2, D2 あたりに「令和6年4月分」などがある
                header_cell_found = False
                for r in range(1, 5):
                    for c in range(1, 10):
                        val = str(sheet.cell(row=r, column=c).value or "")
                        if "令和" in val and "年" in val and "月" in val:
                            m = re.search(r'令和\s*(\d+)\s*年\s*(\d+)\s*月', val.replace('　', ' '))
                            if m:
                                year = 2018 + int(m.group(1))
                                month = int(m.group(2))
                                header_cell_found = True
                                break
                    if header_cell_found: break
                
                if not header_cell_found:
                    # ファイル名から取得 勤務2024.4月分.xlsx
                    m = re.search(r'勤務(\d+)\.(\d+)月', os.path.basename(file_path))
                    if m:
                        year = int(m.group(1))
                        month = int(m.group(2))

                # 12行目から末尾まで
                count = 0
                for row_idx in range(12, 43):
                    day_val = sheet.cell(row=row_idx, column=2).value
                    if day_val is None: continue
                    try:
                        day = int(day_val)
                    except:
                        continue
                        
                    # 出勤判定: 列J(10) または他の列に「開始時間」があるか
                    # A24.4月分シートでは10列目に 10:00:00 
                    start_time = sheet.cell(row=row_idx, column=10).value
                    if start_time:
                        date_str = f"{year}-{month:02d}-{day:02d}"
                        all_data.append({
                            "user_name": user_name,
                            "date": date_str,
                            "is_recorded": True,
                            "source_file": os.path.basename(file_path)
                        })
                        count += 1
                # print(f"  - {user_name}: {count} days")
        except Exception as e:
            print(f"Error processing {file_path}: {e}")
            
    return all_data

folders = [
    '/Users/muratafutoshishi/Library/CloudStorage/GoogleDrive-d.murata@izaya.llc/マイドライブ/経理関係/第２期（2024.3.1-2025.2.28)/勤務表（2024.3.1-2025.2.28)/',
    '/Users/muratafutoshishi/Library/CloudStorage/GoogleDrive-d.murata@izaya.llc/マイドライブ/経理関係/第３期（2025.3.1-2026.2.29)/勤務表（2025.3.1-2026.2.29)/'
]

final_results = []
for f in folders:
    if os.path.exists(f):
        final_results.extend(parse_excel_folder(f))
    else:
        print(f"Folder not found: {f}")

with open('accounting_data.json', 'w', encoding='utf-8') as f:
    json.dump(final_results, f, ensure_ascii=False, indent=2)

print(f"Total records extracted: {len(final_results)}")
print("Results saved to accounting_data.json")
