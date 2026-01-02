
import openpyxl
import os

file_path = '/Users/muratafutoshishi/Library/CloudStorage/GoogleDrive-d.murata@izaya.llc/マイドライブ/経理関係/第２期（2024.3.1-2025.2.28)/勤務表（2024.3.1-2025.2.28)/勤務2024.4月分.xlsx'

wb = openpyxl.load_workbook(file_path, data_only=True)
print(f"File: {os.path.basename(file_path)}")
print(f"Sheets: {wb.sheetnames}")

for name in wb.sheetnames[:5]: # 最初の5シート
    sheet = wb[name]
    print(f"\n--- Sheet: {name} ---")
    # 氏名がどこにあるか探す
    # セル B1-E5 くらいを全部出す
    for r in range(1, 6):
        row_vals = [str(sheet.cell(row=r, column=c).value or "") for c in range(1, 10)]
        print(f"Row {r}: {' | '.join(row_vals)}")
    
    # 12行目付近の構造
    print("Row 11-13:")
    for r in range(11, 14):
        row_vals = [str(sheet.cell(row=r, column=c).value or "") for c in range(1, 15)]
        print(f"Row {r}: {' | '.join(row_vals)}")
