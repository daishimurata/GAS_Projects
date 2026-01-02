
import openpyxl
import os

file_path = '/Users/muratafutoshishi/Library/CloudStorage/GoogleDrive-d.murata@izaya.llc/マイドライブ/経理関係/第２期（2024.3.1-2025.2.28)/勤務表（2024.3.1-2025.2.28)/勤務2024.11月分.xlsx'

wb = openpyxl.load_workbook(file_path, data_only=True)
print(f"File: {os.path.basename(file_path)}")
print(f"All Sheets: {wb.sheetnames}")

# スタッフ関連と思われるシートや、通常とは異なるシートを調査
for name in wb.sheetnames:
    sheet = wb[name]
    # 最初の方の数行を出して中身を推測
    print(f"\n--- Sheet: {name} ---")
    for r in range(1, 4):
        row_vals = [str(sheet.cell(row=r, column=c).value or "") for c in range(1, 10)]
        print(f"Row {r}: {' | '.join(row_vals)}")
