
import openpyxl
import os
import glob
import json
import re
from datetime import datetime
import calendar

def parse_excel_file(file_path):
    all_attendance = []
    print(f"Processing: {os.path.basename(file_path)}")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # 1. 利用者データの抽出 (個別シート)
        for sheet_name in wb.sheetnames:
            user_name = sheet_name.strip()
            if not user_name or user_name in ["合計", "Sheet1", "マスタ", "基本形", "1", "2", "出勤簿", "勤務表", "就労継続支援"]: continue
            
            sheet = wb[sheet_name]
            year, month = guess_year_month(sheet, os.path.basename(file_path))
            
            # 日付の有効範囲を取得 (1日から月末まで)
            _, last_day = calendar.monthrange(year, month)

            for row_idx in range(12, 43):
                day_val = sheet.cell(row=row_idx, column=2).value
                if day_val is None: continue
                try:
                    day = int(day_val)
                    if day < 1 or day > last_day: continue # 不正な日付(0日など)をスキップ
                except:
                    continue
                
                # 出勤判定 (列J=10)
                start_time = sheet.cell(row=row_idx, column=10).value
                if start_time:
                    date_str = f"{year}-{month:02d}-{day:02d}"
                    all_attendance.append({
                        "type": "USER",
                        "user_name": user_name,
                        "date": date_str,
                        "is_recorded": True,
                        "source_file": os.path.basename(file_path)
                    })

        # 2. スタッフデータの抽出 (「出勤簿」シート)
        if "出勤簿" in wb.sheetnames:
            staff_sheet = wb["出勤簿"]
            year, month = guess_year_month(staff_sheet, os.path.basename(file_path))
            _, last_day = calendar.monthrange(year, month)
            
            for r in range(4, 100): # スタッフ数に合わせて調整
                staff_name = staff_sheet.cell(row=r, column=1).value
                if not staff_name or str(staff_name).strip() == "": break
                staff_name = str(staff_name).strip()
                if staff_name in ["氏名", "合計"]: continue
                
                for day in range(1, last_day + 1):
                    col_idx = day + 1
                    status = staff_sheet.cell(row=r, column=col_idx).value
                    if status:
                        date_str = f"{year}-{month:02d}-{day:02d}"
                        all_attendance.append({
                            "type": "STAFF",
                            "user_name": staff_name,
                            "date": date_str,
                            "is_recorded": True,
                            "source_file": os.path.basename(file_path)
                        })

    except Exception as e:
        print(f"Error processing {file_path}: {e}")
            
    return all_attendance

def guess_year_month(sheet, filename):
    year, month = 2024, 4
    header_found = False
    for r in range(1, 5):
        for c in range(1, 15):
            val = str(sheet.cell(row=r, column=c).value or "")
            if "令和" in val or "/" in val:
                m_reiwa = re.search(r'令和\s*(\d+)\s*年\s*(\d+)\s*月', val.replace('　', ' '))
                if m_reiwa:
                    year = 2018 + int(m_reiwa.group(1))
                    month = int(m_reiwa.group(2))
                    header_found = True
                    break
                m_slash = re.search(r'(\d+)/(\d+)', val)
                if m_slash:
                    year = int(m_slash.group(1))
                    month = int(m_slash.group(2))
                    header_found = True
                    break
        if header_found: break
    
    if not header_found:
        m = re.search(r'勤務(\d+)\.(\d+)月', filename)
        if m:
            year, month = int(m.group(1)), int(m.group(2))
            
    return year, month

folders = [
    '/Users/muratafutoshishi/Library/CloudStorage/GoogleDrive-d.murata@izaya.llc/マイドライブ/経理関係/第２期（2024.3.1-2025.2.28)/勤務表（2024.3.1-2025.2.28)/',
    '/Users/muratafutoshishi/Library/CloudStorage/GoogleDrive-d.murata@izaya.llc/マイドライブ/経理関係/第３期（2025.3.1-2026.2.29)/勤務表（2025.3.1-2026.2.29)/'
]

final_results = []
for folder in folders:
    if os.path.exists(folder):
        files = glob.glob(os.path.join(folder, "勤務*.xlsx"))
        files.sort()
        for f in files:
            final_results.extend(parse_excel_file(f))

with open('attendance_combined_data.json', 'w', encoding='utf-8') as f:
    json.dump(final_results, f, ensure_ascii=False, indent=2)

print(f"Total combined records (User & Staff): {len(final_results)}")
print("Results saved to attendance_combined_data.json")
